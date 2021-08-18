from openpyxl import load_workbook
from re import sub
from os import listdir
from pathlib import Path

def logger(filename=None, *args, **kwargs):
    #print(*args, **kwargs)
    if filename is not None:
        with open(filename + '.logs', 'a', encoding="utf-8") as file:
            try: 
                file.write(''.join(args))
            except UnicodeEncodeError:
                print("Line contains UNICODE chars")
                


def change_address_to_absolute(wbname, workbook):
    logger(wbname, f"\n[***] Reading file: {wbname}\n")
    # Enumerate worksheets in loaded workbook
    for ws in workbook:
        logger(wbname, f"[*] Reading sheet: {ws.title}, size: {ws.max_row}x{ws.max_column}\n\n")
        for index, row in enumerate(ws):
            for record in row:
                # Replace only equations
                if str(record.value).startswith('='):
                    logger(wbname, f"[*] Old value: {record.value}")
                    # Insert middle $ (absolute address mark) if there is none
                    middle_absolute = sub(r'([a-zA-Z]+)(?!\$)(\d+)', r'\1$\2', record.value)
                    # Insert beginning $ if there is none (negative lookbehind)
                    beg_absolute = sub(r'(?<!\$)([a-zA-Z]+\$\d+)', r'$\1', middle_absolute)
                    logger(wbname, f"[*] New value: {beg_absolute}", "\n---\n")

    # Save the file
    print(f"[*]Saving workbook: {wbname} to file, {Path(wbname).stem + '-matched' + Path(wbname).suffix}")
    workbook.save(Path(wbname).stem + '-matched' + Path(wbname).suffix)

#

print("[*] Reading Excel workbooks in current directory")
for file in listdir():
    if file.endswith('.xltx'):
        print("[**] Reading Workbook: ", file)
        wb = load_workbook(file)
        change_address_to_absolute(file, wb)
